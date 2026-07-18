"""Lógica de negocio del dominio de finanzas.

Reglas de integridad del dinero:
- gasto:        resta `monto` al saldo de la cuenta.
- ingreso:      suma `monto` al saldo de la cuenta.
- transferencia: mueve `monto` de la cuenta origen a la cuenta destino.
Al eliminar un movimiento se revierte su efecto sobre los saldos.

Los errores de validación (cuenta inexistente, transferencia mal formada, etc.)
se señalan con ValueError y el router los traduce a 400.
"""

import calendar
import uuid
from datetime import date
from decimal import Decimal
from io import BytesIO

from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

from app.models.finances import Account, Budget, Category, Transaction
from app.schemas.finances import (
    AccountCreate,
    AccountUpdate,
    BudgetCreate,
    BudgetUpdate,
    CategoryCreate,
    CategoryUpdate,
    TransactionCreate,
)

GASTO = "gasto"
INGRESO = "ingreso"
TRANSFERENCIA = "transferencia"


# --- Cuentas ---


def create_account(db: Session, data: AccountCreate) -> Account:
    account = Account(nombre=data.nombre, tipo=data.tipo, saldo=data.saldo_inicial)
    db.add(account)
    db.commit()
    db.refresh(account)
    return account


def get_account(db: Session, account_id: uuid.UUID) -> Account | None:
    return db.get(Account, account_id)


def list_accounts(db: Session) -> list[Account]:
    return list(db.execute(select(Account).order_by(Account.nombre)).scalars().all())


def update_account(
    db: Session, account_id: uuid.UUID, data: AccountUpdate
) -> Account | None:
    account = db.get(Account, account_id)
    if account is None:
        return None
    for campo, valor in data.model_dump(exclude_unset=True).items():
        setattr(account, campo, valor)
    db.commit()
    db.refresh(account)
    return account


# --- Categorías ---


def create_category(db: Session, data: CategoryCreate) -> Category:
    if db.execute(
        select(Category).where(func.lower(Category.nombre) == data.nombre.lower())
    ).scalar_one_or_none():
        raise ValueError("Ya existe una categoría con ese nombre")
    category = Category(nombre=data.nombre, activa=data.activa)
    db.add(category)
    db.commit()
    db.refresh(category)
    return category


def list_categories(db: Session, solo_activas: bool = False) -> list[Category]:
    stmt = select(Category)
    if solo_activas:
        stmt = stmt.where(Category.activa.is_(True))
    return list(db.execute(stmt.order_by(Category.nombre)).scalars().all())


def update_category(
    db: Session, category_id: uuid.UUID, data: CategoryUpdate
) -> Category | None:
    category = db.get(Category, category_id)
    if category is None:
        return None
    for campo, valor in data.model_dump(exclude_unset=True).items():
        setattr(category, campo, valor)
    db.commit()
    db.refresh(category)
    return category


# --- Movimientos ---


def _mover_saldos(
    tipo: str,
    monto: Decimal,
    origen: Account,
    destino: Account | None,
    signo: int,
) -> None:
    """Aplica (signo=+1) o revierte (signo=-1) el efecto en los saldos."""
    delta = monto * signo
    if tipo == GASTO:
        origen.saldo -= delta
    elif tipo == INGRESO:
        origen.saldo += delta
    elif tipo == TRANSFERENCIA:
        origen.saldo -= delta
        assert destino is not None
        destino.saldo += delta


def create_transaction(db: Session, data: TransactionCreate) -> Transaction:
    origen = db.get(Account, data.account_id)
    if origen is None:
        raise ValueError("La cuenta de origen no existe")

    destino: Account | None = None
    tipo = data.tipo.value

    if tipo == TRANSFERENCIA:
        if data.cuenta_destino_id is None:
            raise ValueError("Una transferencia requiere cuenta de destino")
        if data.cuenta_destino_id == data.account_id:
            raise ValueError("La cuenta de destino debe ser distinta de la de origen")
        if data.category_id is not None:
            raise ValueError("Una transferencia no lleva categoría")
        destino = db.get(Account, data.cuenta_destino_id)
        if destino is None:
            raise ValueError("La cuenta de destino no existe")
    else:  # gasto / ingreso
        if data.cuenta_destino_id is not None:
            raise ValueError("Solo las transferencias llevan cuenta de destino")
        if data.category_id is None:
            raise ValueError("Un gasto o ingreso requiere categoría")
        if db.get(Category, data.category_id) is None:
            raise ValueError("La categoría no existe")

    tx = Transaction(
        tipo=tipo,
        monto=data.monto,
        account_id=data.account_id,
        cuenta_destino_id=data.cuenta_destino_id,
        category_id=data.category_id,
        fecha=data.fecha or date.today(),
        nota=data.nota,
    )
    _mover_saldos(tipo, data.monto, origen, destino, signo=1)
    db.add(tx)
    db.commit()
    db.refresh(tx)
    return tx


def get_transaction(db: Session, tx_id: uuid.UUID) -> Transaction | None:
    return db.get(Transaction, tx_id)


def update_transaction(db: Session, tx_id: uuid.UUID, data) -> Transaction | None:
    """Edita un movimiento sin cambiar su tipo. Solo se tocan los campos
    enviados; el saldo se mantiene coherente: se revierte el efecto anterior
    y se aplica el nuevo. Errores de validación via ValueError (-> 400)."""
    tx = db.get(Transaction, tx_id)
    if tx is None:
        return None

    cambios = data.model_dump(exclude_unset=True)
    tipo = tx.tipo  # inmutable

    nuevo_monto = cambios.get("monto", tx.monto)
    nueva_cuenta = cambios.get("account_id", tx.account_id)
    nueva_categoria = cambios.get("category_id", tx.category_id)
    nuevo_destino = cambios.get("cuenta_destino_id", tx.cuenta_destino_id)
    nueva_fecha = cambios.get("fecha", tx.fecha)
    nueva_nota = cambios.get("nota", tx.nota)

    # --- validar la combinación resultante según el tipo ---
    origen_new = db.get(Account, nueva_cuenta)
    if origen_new is None:
        raise ValueError("La cuenta de origen no existe")

    destino_new: Account | None = None
    if tipo == TRANSFERENCIA:
        if nuevo_destino is None:
            raise ValueError("Una transferencia requiere cuenta de destino")
        if nuevo_destino == nueva_cuenta:
            raise ValueError("La cuenta de destino debe ser distinta de la de origen")
        if nueva_categoria is not None:
            raise ValueError("Una transferencia no lleva categoría")
        destino_new = db.get(Account, nuevo_destino)
        if destino_new is None:
            raise ValueError("La cuenta de destino no existe")
    else:  # gasto / ingreso
        if nuevo_destino is not None:
            raise ValueError("Solo las transferencias llevan cuenta de destino")
        if nueva_categoria is None:
            raise ValueError("Un gasto o ingreso requiere categoría")
        if db.get(Category, nueva_categoria) is None:
            raise ValueError("La categoría no existe")

    # --- revertir el efecto anterior (con los valores viejos) ---
    origen_old = db.get(Account, tx.account_id)
    destino_old = (
        db.get(Account, tx.cuenta_destino_id) if tx.cuenta_destino_id else None
    )
    if origen_old is not None:
        _mover_saldos(tipo, tx.monto, origen_old, destino_old, signo=-1)

    # --- aplicar los valores nuevos y su efecto ---
    tx.monto = nuevo_monto
    tx.account_id = nueva_cuenta
    tx.category_id = nueva_categoria
    tx.cuenta_destino_id = nuevo_destino
    tx.fecha = nueva_fecha
    tx.nota = nueva_nota
    _mover_saldos(tipo, nuevo_monto, origen_new, destino_new, signo=1)

    db.commit()
    db.refresh(tx)
    return tx


def list_transactions(
    db: Session,
    account_id: uuid.UUID | None = None,
    tipo: str | None = None,
    category_id: uuid.UUID | None = None,
    desde: date | None = None,
    hasta: date | None = None,
) -> list[Transaction]:
    stmt = select(Transaction)
    if account_id is not None:
        stmt = stmt.where(Transaction.account_id == account_id)
    if tipo is not None:
        stmt = stmt.where(Transaction.tipo == tipo)
    if category_id is not None:
        stmt = stmt.where(Transaction.category_id == category_id)
    if desde is not None:
        stmt = stmt.where(Transaction.fecha >= desde)
    if hasta is not None:
        stmt = stmt.where(Transaction.fecha <= hasta)
    stmt = stmt.order_by(Transaction.fecha.desc())
    return list(db.execute(stmt).scalars().all())


def delete_transaction(db: Session, tx_id: uuid.UUID) -> bool:
    """Elimina un movimiento y revierte su efecto sobre los saldos."""
    tx = db.get(Transaction, tx_id)
    if tx is None:
        return False
    origen = db.get(Account, tx.account_id)
    destino = (
        db.get(Account, tx.cuenta_destino_id)
        if tx.cuenta_destino_id is not None
        else None
    )
    if origen is not None:
        _mover_saldos(tx.tipo, tx.monto, origen, destino, signo=-1)
    db.delete(tx)
    db.commit()
    return True


# --- Reportes y presupuestos ---


def _rango_mes(anio: int, mes: int) -> tuple[date, date]:
    ultimo = calendar.monthrange(anio, mes)[1]
    return date(anio, mes, 1), date(anio, mes, ultimo)


def reporte_mensual(
    db: Session, anio: int, mes: int
) -> tuple[Decimal, list[tuple[uuid.UUID | None, str, Decimal]]]:
    """Total de gastos del mes y desglose por categoría.

    Solo cuenta `tipo=gasto` (las transferencias no son gasto real)."""
    inicio, fin = _rango_mes(anio, mes)
    stmt = (
        select(
            Transaction.category_id,
            func.coalesce(Category.nombre, "(sin categoría)"),
            func.sum(Transaction.monto),
        )
        .join(Category, Category.id == Transaction.category_id, isouter=True)
        .where(
            Transaction.tipo == GASTO,
            Transaction.fecha >= inicio,
            Transaction.fecha <= fin,
        )
        .group_by(Transaction.category_id, Category.nombre)
        .order_by(func.sum(Transaction.monto).desc())
    )
    filas = db.execute(stmt).all()
    por_categoria = [(cid, nombre, total) for cid, nombre, total in filas]
    total = sum((t for _, _, t in por_categoria), Decimal("0"))
    return total, por_categoria


def _gasto_del_mes(
    db: Session, category_id: uuid.UUID | None, inicio: date, fin: date
) -> Decimal:
    stmt = select(func.coalesce(func.sum(Transaction.monto), 0)).where(
        Transaction.tipo == GASTO,
        Transaction.fecha >= inicio,
        Transaction.fecha <= fin,
    )
    if category_id is not None:
        stmt = stmt.where(Transaction.category_id == category_id)
    return Decimal(db.execute(stmt).scalar_one())


def create_budget(db: Session, data: BudgetCreate) -> Budget:
    if data.category_id is not None and db.get(Category, data.category_id) is None:
        raise ValueError("La categoría no existe")
    budget = Budget(
        category_id=data.category_id, tope=data.tope, periodo=data.periodo
    )
    db.add(budget)
    db.commit()
    db.refresh(budget)
    return budget


def get_budget(db: Session, budget_id: uuid.UUID) -> Budget | None:
    return db.get(Budget, budget_id)


def list_budgets(db: Session) -> list[Budget]:
    return list(db.execute(select(Budget)).scalars().all())


def update_budget(
    db: Session, budget_id: uuid.UUID, data: BudgetUpdate
) -> Budget | None:
    budget = db.get(Budget, budget_id)
    if budget is None:
        return None
    for campo, valor in data.model_dump(exclude_unset=True).items():
        setattr(budget, campo, valor)
    db.commit()
    db.refresh(budget)
    return budget


def delete_budget(db: Session, budget_id: uuid.UUID) -> bool:
    budget = db.get(Budget, budget_id)
    if budget is None:
        return False
    db.delete(budget)
    db.commit()
    return True


def budget_progress(
    db: Session, budget_id: uuid.UUID, anio: int, mes: int
) -> tuple[Budget, Decimal] | None:
    """Devuelve (presupuesto, gastado_en_el_mes). Restante y % los arma el
    router. Un presupuesto sin categoría suma todos los gastos del mes."""
    budget = db.get(Budget, budget_id)
    if budget is None:
        return None
    inicio, fin = _rango_mes(anio, mes)
    gastado = _gasto_del_mes(db, budget.category_id, inicio, fin)
    return budget, gastado


# --- Exportación a Excel ---

_MONEY_FMT = '"$"#,##0'
_DATE_FMT = "yyyy-mm-dd"
_HEADER_COLOR = "2F6F6D"  # teal de la marca


def _tx_export_query(
    db: Session,
    account_id: uuid.UUID | None,
    tipo: str | None,
    category_id: uuid.UUID | None,
    desde: date | None,
    hasta: date | None,
) -> list[Transaction]:
    """Como list_transactions, pero al filtrar por cuenta incluye las
    transferencias donde la cuenta es origen O destino (para que el libro
    coincida con el detalle de cuenta de la web)."""
    stmt = select(Transaction)
    if account_id is not None:
        stmt = stmt.where(
            or_(
                Transaction.account_id == account_id,
                Transaction.cuenta_destino_id == account_id,
            )
        )
    if tipo is not None:
        stmt = stmt.where(Transaction.tipo == tipo)
    if category_id is not None:
        stmt = stmt.where(Transaction.category_id == category_id)
    if desde is not None:
        stmt = stmt.where(Transaction.fecha >= desde)
    if hasta is not None:
        stmt = stmt.where(Transaction.fecha <= hasta)
    return list(db.execute(stmt.order_by(Transaction.fecha.desc())).scalars().all())


def export_transactions_xlsx(
    db: Session,
    account_id: uuid.UUID | None = None,
    tipo: str | None = None,
    category_id: uuid.UUID | None = None,
    desde: date | None = None,
    hasta: date | None = None,
) -> bytes:
    """Arma un .xlsx con los movimientos filtrados (hoja Movimientos) y sus
    totales (hoja Resumen). Devuelve los bytes del libro."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    txs = _tx_export_query(db, account_id, tipo, category_id, desde, hasta)
    cuentas = {a.id: a.nombre for a in list_accounts(db)}
    categorias = {c.id: c.nombre for c in list_categories(db)}

    def nom_cuenta(cid: uuid.UUID | None) -> str:
        return cuentas.get(cid, "—") if cid else ""

    def nom_categoria(cid: uuid.UUID | None) -> str:
        return categorias.get(cid, "—") if cid else "(sin categoría)"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=_HEADER_COLOR)
    bold = Font(bold=True)

    wb = Workbook()

    # --- Hoja Movimientos ---
    ws = wb.active
    ws.title = "Movimientos"
    encabezados = ["Fecha", "Tipo", "Monto", "Cuenta", "Cuenta destino", "Categoría", "Nota"]
    ws.append(encabezados)
    for col in range(1, len(encabezados) + 1):
        celda = ws.cell(row=1, column=col)
        celda.font = header_font
        celda.fill = header_fill
    for t in txs:
        ws.append([
            t.fecha,
            t.tipo,
            float(t.monto),
            nom_cuenta(t.account_id),
            nom_cuenta(t.cuenta_destino_id) if t.tipo == TRANSFERENCIA else "",
            nom_categoria(t.category_id) if t.tipo != TRANSFERENCIA else "",
            t.nota or "",
        ])
    for fila in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        fila[0].number_format = _DATE_FMT
    for fila in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        fila[0].number_format = _MONEY_FMT
    for i, ancho in enumerate([12, 14, 14, 16, 16, 18, 40], start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.freeze_panes = "A2"

    # --- Hoja Resumen ---
    rs = wb.create_sheet("Resumen")
    ingresos = sum((t.monto for t in txs if t.tipo == INGRESO), Decimal("0"))
    gastos = sum((t.monto for t in txs if t.tipo == GASTO), Decimal("0"))

    def money(celda: str, valor: Decimal, negrita: bool = False) -> None:
        rs[celda] = float(valor)
        rs[celda].number_format = _MONEY_FMT
        if negrita:
            rs[celda].font = bold

    rs["A1"] = "Resumen de finanzas"
    rs["A1"].font = bold
    rs["A2"] = "Período"
    rs["B2"] = f"{desde or '—'} a {hasta or 'hoy'}"
    rs["A3"] = "Ingresos"; money("B3", ingresos)
    rs["A4"] = "Gastos"; money("B4", gastos)
    rs["A5"] = "Diferencia"; rs["A5"].font = bold; money("B5", ingresos - gastos, negrita=True)

    def encabezar(fila: int, titulos: list[str]) -> None:
        for col, txt in zip("ABCD", titulos):
            rs[f"{col}{fila}"] = txt
            rs[f"{col}{fila}"].font = bold

    # Por categoría
    fila = 7
    rs[f"A{fila}"] = "Por categoría"; rs[f"A{fila}"].font = bold
    fila += 1
    encabezar(fila, ["Categoría", "Ingresos", "Gastos", "Neto"])
    fila += 1
    por_cat: dict[str, list[Decimal]] = {}
    for t in txs:
        if t.tipo not in (GASTO, INGRESO):
            continue
        acc = por_cat.setdefault(nom_categoria(t.category_id), [Decimal("0"), Decimal("0")])
        acc[0 if t.tipo == INGRESO else 1] += t.monto
    for nom, (ing, gas) in sorted(por_cat.items(), key=lambda kv: abs(kv[1][0] - kv[1][1]), reverse=True):
        rs[f"A{fila}"] = nom
        money(f"B{fila}", ing); money(f"C{fila}", gas); money(f"D{fila}", ing - gas)
        fila += 1

    # Por cuenta (entradas/salidas/neto, incluyendo transferencias)
    fila += 1
    rs[f"A{fila}"] = "Por cuenta"; rs[f"A{fila}"].font = bold
    fila += 1
    encabezar(fila, ["Cuenta", "Entradas", "Salidas", "Neto"])
    fila += 1
    por_cuenta: dict[str, list[Decimal]] = {}
    for t in txs:
        origen = por_cuenta.setdefault(nom_cuenta(t.account_id), [Decimal("0"), Decimal("0")])
        if t.tipo == INGRESO:
            origen[0] += t.monto
        else:  # gasto o transferencia: sale de la cuenta origen
            origen[1] += t.monto
        if t.tipo == TRANSFERENCIA and t.cuenta_destino_id:
            destino = por_cuenta.setdefault(nom_cuenta(t.cuenta_destino_id), [Decimal("0"), Decimal("0")])
            destino[0] += t.monto  # entra a la cuenta destino
    for nom, (ent, sal) in sorted(por_cuenta.items()):
        rs[f"A{fila}"] = nom
        money(f"B{fila}", ent); money(f"C{fila}", sal); money(f"D{fila}", ent - sal)
        fila += 1

    for col, ancho in zip("ABCD", [22, 14, 14, 14]):
        rs.column_dimensions[col].width = ancho

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
